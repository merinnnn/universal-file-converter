import asyncio
import csv as csv_module
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .base import Converter, ConversionError
from .registry import register
from .documents import LibreOfficeConverter

class CsvToXlsxConverter(Converter):
    async def convert(self, input_path: Path, output_path: Path) -> None:
        await asyncio.to_thread(self._convert_sync, input_path, output_path)

    def _convert_sync(self, input_path: Path, output_path: Path) -> None:
        try:
            wb = Workbook()
            ws = wb.active
            with open(input_path, newline="", encoding="utf-8-sig") as f:
                for row in csv_module.reader(f):
                    ws.append(row)
            wb.save(output_path)
        except Exception as e:
            raise ConversionError(f"csv-to-xlsx failed: {e}")

class XlsxToCsvConverter(Converter):
    async def convert(self, input_path: Path, output_path: Path) -> None:
        await asyncio.to_thread(self._convert_sync, input_path, output_path)

    def _convert_sync(self, input_path: Path, output_path: Path) -> None:
        try:
            wb = load_workbook(input_path, data_only=True)
            ws = wb.active
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv_module.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if c is None else c for c in row])
        except Exception as e:
            raise ConversionError(f"xlsx-to-csv failed: {e}")

register("csv", "xlsx")(CsvToXlsxConverter)
register("xlsx", "csv")(XlsxToCsvConverter)

# xlsx <-> ods via the already-proven generic LibreOffice engine
register("xlsx", "ods")(LibreOfficeConverter)
register("ods", "xlsx")(LibreOfficeConverter)